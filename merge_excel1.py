import os
import glob
import openpyxl

def new_file(PATH,text):
    wbook=openpyxl.Workbook()
    wsheet=wbook.active
    ws1=wbook.create_sheet('生产计划',0)
    ws2=wbook.create_sheet('停机时间汇总',1)
    wbook.save(text)
    xlsx_files = get_all_xlsx_files(os.path.expanduser (PATH)) #定义变量xlsx_files为get_all_xlsx_files函数，指定参数为指定目录
    wb=openpyxl.load_workbook(xlsx_files[0])
    ws3=wb.worksheets[1]
    for filename in xlsx_files[1:2]:
        workbook=openpyxl.load_workbook(filename)
        sheet=workbook.worksheets[1]
        for row in sheet.iter_rows(min_row=1,max_row=3): #遍历其他文件，忽略首行内容
                values = [cell.value for cell in row] #循环获取单元格的值
                ws3.append(values)#将值依次添加末尾
    wb.save(text)

def merge_xlsx_files(xlsx_files,key):  #定义函数合并xlsx文件
    wb = openpyxl.load_workbook(xlsx_files[0]) #调用openpyxl模块load_workbook函数
    ws = wb.worksheets[1]                             #获取活跃的Worksheet

    for filename in xlsx_files[0:]:            #循环xlsx_files参数，获取第一个工作表（只有一个）
        if key in filename:
            workbook = openpyxl.load_workbook(filename)  #调用函数
            sheet = workbook.worksheets[1]                  #获取活跃的表格
            for row in sheet.iter_rows(min_row=4,max_row=sheet.max_row+1): #遍历其他文件，忽略首行内容
                values = [cell.value for cell in row] #循环获取单元格的值
                ws.append(values)#将值依次添加末尾
    print(xlsx_files[0])
    return wb                                     #返回

def get_all_xlsx_files(path):                        #定义获取所有xlsx文件
    xlsx_files = glob.glob(os.path.join(path,'*.xlsx') )#采用glob方法指定路径下所有.xlsx的文件
    #sorted(xlsx_files,key=str.lower)                     #按照关键字字符串小写排序
    return xlsx_files

def merge(text1,text2,PATH):
    xlsx_files = get_all_xlsx_files(os.path.expanduser (PATH)) #定义变量xlsx_files为get_all_xlsx_files函数，指定参数为指定目录
    wb = merge_xlsx_files(xlsx_files,text1)                            #定义wb为merge_xlsx_files函数，指定参数为遍历
    sht=wb.worksheets[1]
    j=4
    for j in sht.merged_cells.ranges[18:]:
        r1,r2,c1,c2=j.min_row,j.max_row,j.min_col,j.max_col
        sht.unmerge_cells(start_row=r1,end_row=r2,start_column=c1,end_column=c2)
        print(sht.merged_cell_ranges)
    i=4
    while i<=sht.max_row:
        if sht.cell(row=i,column=2).value is None:
            sht.delete_rows(idx=i)
        else:
            i+=1
    k=sht.max_row
    cell1=[]
    val=0
    for row in range(4,k+1):
        if sht.cell(row=row,column=9).value is not None:
            val=val+sht.cell(row=row,column=9).value
            cell1.append(val)
            m=len(cell1)
            SUM=cell1[m-1]
            sht["S4"]=SUM
    wb.save(text2+".xlsx")                                  #save方法将汇总表保存到merged_form.xlsx

def main():                                        #定义主函数
    PATH=input("path:")
    new_file(PATH,PATH+'\\'"酸轧停机汇总.xlsx")
    merge("甲","甲班停机汇总",PATH)
    merge("乙","乙班停机汇总",PATH)
    merge("丙","丙班停机汇总",PATH)
if __name__ =='__main__':
    main()